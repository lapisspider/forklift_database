"""Import the 'Model Numbers' spreadsheet (Sheet1) as KITS.

Each spreadsheet row maps an SSC kit SKU (174-...) to the forklift model(s)/series
it fits. We create one Kit per unique SKU:
  - sku    <- column B (174-...)          (required)
  - name   <- column A when it's an SSC kit-series name (Apollo, ...)  (optional)
  - notes  <- OEM + the raw "fits" model text + caveats, aggregated across every
              row with that SKU. The enrichment step (enrich_forklifts.py) reads
              this to create individual Forklift models and link them.

Forklifts are NOT created here — they come from enrich_forklifts.py (web-lookup,
one entry per individual model, each with a spec sheet).

WARNING: this WIPES the forklifts, kits, and kit_forklift tables (authorized reset).

Usage:  python import_models.py "path\\to\\Model Numbers - 2026 (1).xlsx"
"""
import re
import sys

from openpyxl import load_workbook

from app.database import SessionLocal, engine, init_db
from app.models import Forklift, Kit, kit_forklift

OEM_HEADERS = {
    "cat-yellow": "CAT (Yellow)", "byd": "BYD", "ottowa": "Ottawa",
    "nissan/unicarrier/logisnext": "Nissan/Unicarrier/Logisnext", "toyota": "Toyota",
    "raymond": "Raymond", "clark": "Clark", "crown": "Crown", "daewoo": "Daewoo",
    "bobcat/doosan": "Bobcat/Doosan", "hyster": "Hyster", "yale": "Yale",
    "hundai": "Hyundai", "jungheinrich": "Jungheinrich", "komatsu": "Komatsu",
    "drexel": "Drexel", "big joe": "Big Joe", "heli": "Heli", "yamaha": "Yamaha",
    "taylor dunn": "Taylor Dunn", "club car": "Club Car", "textron": "Textron",
    "stewart & stevenson": "Stewart & Stevenson",
}
KIT_SERIES_NAMES = {
    "apollo", "jupiter", "atlas", "magellan", "mars", "mercury", "luna",
    "kibo", "helios",
}
REPORT = "no_sku_report.txt"


def clean(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).replace("\xa0", " ")).strip()


def section_for(text: str) -> str | None:
    t = clean(text).lower()
    if t in OEM_HEADERS:
        return OEM_HEADERS[t]
    if t.startswith("linde"):
        return "Linde"
    return None


def main(path: str) -> None:
    init_db()
    wb = load_workbook(path, data_only=True)
    ws = wb["Sheet1"]

    db = SessionLocal()
    # Authorized reset.
    db.execute(kit_forklift.delete())
    db.query(Kit).delete()
    db.query(Forklift).delete()
    db.commit()

    kits: dict[str, dict] = {}   # sku -> {name, oem, fits[], caveats[]}
    no_sku: list[str] = []
    oem = "CAT/Mitsubishi"        # the block before the first OEM header

    for row in ws.iter_rows(min_row=4, values_only=True):
        a, b, c = clean(row[0]), clean(row[1]), clean(row[2])
        rest = [clean(x) for x in row[3:] if clean(x)]

        sec = section_for(c)
        if sec and not b.startswith("174-"):
            oem = sec
            continue
        if not c:
            continue
        # Continuation/commentary lines: no SKU, long prose.
        if not b and len(c) > 70:
            continue

        if b.startswith("174-"):
            e = kits.setdefault(b, {"name": None, "oem": oem, "fits": [], "caveats": []})
            if a:
                first = a.split()[0].split("-")[0].strip().lower()
                if first in KIT_SERIES_NAMES and not e["name"]:
                    e["name"] = a
                elif not e["name"] and first not in KIT_SERIES_NAMES:
                    e["caveats"].append(f"row-label: {a}")
            e["fits"].append(c)
            e["caveats"].extend(rest)
        else:
            note = f"{oem} | {c}" + (f" | {a}" if a else "")
            no_sku.append(note)

    for sku, e in kits.items():
        notes = f"OEM: {e['oem']}\nFits: " + " | ".join(e["fits"])
        if e["caveats"]:
            notes += "\nCaveats: " + " | ".join(e["caveats"])
        db.add(Kit(sku=sku, name=e["name"], notes=notes))
    db.commit()

    with open(REPORT, "w", encoding="utf-8") as fh:
        fh.write("Rows with a forklift model but NO kit SKU (handle manually):\n\n")
        fh.write("\n".join(no_sku))

    total = db.query(Kit).count()
    db.close()
    print(f"Imported {total} kits (unique SKUs). "
          f"{len(no_sku)} model rows had no SKU -> {REPORT}.")


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else
         r"C:\Users\Jacob Barnhill\Downloads\Model Numbers - 2026 (1).xlsx")
